import openpyxl


# -说明[xlrd/xlwt在读写方面的速度都要优于openpyxl，但无法生成.xlsx文件且对大文件支持不好]

def test_read():
    # 创建对象
    wb = openpyxl.load_workbook('test_read.xlsx')

    # 获取所有sheet name
    sheet_names = wb.sheetnames
    print('sheet_names:', sheet_names)

    # 获取某sheet对象的集中方式
    sheet1 = wb['第一张表']
    sheet2 = wb[sheet_names[1]]
    sheet3 = wb.worksheets[2]
    sheet_act = wb.active  # 当前活跃的worksheet[当前打开的sheet]
    print('sheet1:', sheet1, '|| sheet2:', sheet2, '|| sheet3:', sheet3, '|| sheet_act:', sheet_act)

    # 获取sheet的属性[表名、行数、列数]
    title = sheet1.title
    max_row = sheet1.max_row
    max_column = sheet1.max_column
    print('title:', title, '|| max_row:', max_row, '|| max_column:', max_column)

    # 按行或列获取sheet中的数据[无数据为None]
    print('按先行后列读取sheet数据↓')
    for row in sheet1.rows:  # row为由元组包含的每一行数据
        for cell in row:
            print(cell.value)
        print(row, '\n')

    print('按先列后行读取sheet数据↓')
    for column in sheet1.columns:
        for cell in column:
            print(cell.value)
        print(column, '\n')

    # 获取特定行或列的数据
    print('获取特定行数据↓')
    for cell in list(sheet1.rows)[2]:
        print(cell.value)

    print('获取特定列数据↓')
    for cell in list(sheet1.columns)[2]:
        print(cell.value)

    # 获取某一块数据
    print('\n获取某一块数据-方式一↓')
    for rows in list(sheet1.rows)[1:3]:
        for cell in rows[1:3]:
            print(cell.value)  # 2-3行2-3列
        print()

    print('\n获取某一块数据-方式二↓')
    for i in range(2, 4):
        for j in range(2, 5):
            print(sheet1.cell(row=i, column=j).value)  # 2-3行2-3-4列
        print()

    # 获取某一单元格数据的两种方式
    value_b3 = sheet1['B3'].value
    value_2_3 = sheet1.cell(row=2, column=3).value
    print('value_b3:', value_b3, '|| value_2_3:', value_2_3)


def test_write():
    # 创建对象
    wb = openpyxl.Workbook()
    # 获取当前活跃的sheet
    sheet1 = wb.active
    sheet1.title = 'write1'
    # 创建sheet
    sheet2 = wb.create_sheet()
    sheet2.title = 'write2'

    # 准备写入的数据
    project = ['各省市', '工资性收入', '家庭经营纯收入', '财产性收入', '转移性收入', '食品', '衣着', '居住', '家庭设备及服务', '交通和通讯', '文教、娱乐用品及服务', '医疗保健', '其他商品及服务']
    province = ['北京市', '天津市', '河北省', '山西省', '内蒙古自治区', '辽宁省', '吉林省', '黑龙江省', '上海市', '江苏省', '浙江省', '安徽省', '福建省', '江西省', '山东省', '河南省', '湖北省', '湖南省', '广东省', '广西壮族自治区', '海南省', '重庆市', '四川省', '贵州省', '云南省', '西藏自治区', '陕西省', '甘肃省', '青海省', '宁夏回族自治区', '新疆维吾尔自治区']
    income = ['5047.4', '3247.9', '1514.7', '1374.3', '590.7', '1499.5', '605.1', '654.9', '6686.0', '3104.8', '3575.1', '1184.1', '1855.5', '1441.3', '1671.5', '1022.7', '1199.2', '1449.6', '2906.2', '972.3', '555.7', '1309.9', '1219.5', '715.5', '441.8', '568.4', '848.3', '637.4', '653.3', '823.1', '254.1']

    # 写第一行[行列号均从1开始计数]
    for i in range(len(project)):
        sheet2.cell(1, i + 1, project[i])

    for i in range(len(province)):
        sheet2.cell(i + 2, 1, province[i])

    for i in range(len(income)):
        sheet2.cell(i + 2, 2, income[i])

    # 保存文件
    wb.save(filename='test_write.xlsx')


def test_update():
    wb = openpyxl.load_workbook('test_write.xlsx')
    sheet = wb.worksheets[1]

    # 插入一行数据[不太规则为了看操作手法加的]
    max_row = sheet.max_row
    sheet.insert_rows(max_row)
    print('max-', sheet.max_row)
    for i, column in enumerate(sheet.columns):
        if i == 0:
            column[max_row].value = 'column start. 1'
        else:
            column[max_row].value = i + 1

    # 插入一列数据
    sheet.insert_cols(1)
    for i, row in enumerate(sheet.rows):
        if i == 0:
            row[0].value = 'raw no.'
        else:
            row[0].value = i

    # 修改特定单元格的两种方式
    sheet.cell(2, 5, '仔')
    sheet['B3'] = '25仔'

    # 添加行的另一种方式
    taiwan = [36, '中国台湾']
    sheet.append(taiwan)

    wb.save(filename='test_write.xlsx')


if __name__ == '__main__':
    # test_read()
    # test_write()
    # test_update()
    pass
